import openpyxl as xl
from openpyxl.chart import BarChart,Reference
import matplotlib.pyplot as plt
import math

wb = xl.load_workbook('Activity_1_Data.xlsx')
sheet = wb['Sheet1']
#cell = sheet['a1']
#cell = sheet.cell(2, 1)
#print(cell.value)
#print(sheet.max_row)

x = []
y = []

for i in range(sheet.max_row):
    v = sheet.cell(i+1,2)
    val = v.value
    v1 = sheet.cell(i+1,3)
    val1 = v1.value
    x.append(val)
    y.append(val1)
print(f'x = {x}')
print(f'y = {y}')

sum = 0
for j in range(170):
    sum = sum + y[j]

mean = sum/170
print(f'mean = {mean}')

nby2 = y[85]
nby21 = y[86]
median = (nby2+nby21)/2
print(f'median = {median}')

mode = max(y, key=y.count)
print(f'mode = {mode}')

y2 = []
for k in range(170):
    v3 = y[i]-mean
    v4 = v3**2
    y2.append(v4)
print(f'y2 = {y2}')

sum2 = 0
for l in range(170):
    sum2 = sum2 + y2[l]
variance = sum2/170
print(f'variance = {variance}')

standard_deviation = math.sqrt(variance)
print(f'standard_deviation = {standard_deviation}')


sumx = 0
for j in range(170):
    sumx = sumx + x[j]

meanx = sumx/170
print(f'meanx = {meanx}')

nby2x = x[85]
nby21x = x[86]
medianx = (nby2x+nby21x)/2
print(f'medianx = {medianx}')

modex = max(x, key=x.count)
print(f'modex = {modex}')

x2 = []
for k in range(170):
    v3x = x[i]-mean
    v4x = v3x**2
    x2.append(v4x)
print(f'x2 = {x2}')

sum2x = 0
for l in range(170):
    sum2x = sum2x + x2[l]
variancex = sum2x/170
print(f'variancex = {variancex}')

standard_deviationx = math.sqrt(variancex)
print(f'standard_deviationx = {standard_deviationx}')

new_x_pos = round(modex)
new_y_pos = round(mode)
print(f'new base station = ({new_x_pos},{new_y_pos})')

values = Reference(sheet,min_row = 1,max_row = sheet.max_row,min_col = 2,max_col = 3 )
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')


wb.save('activity1.xlsx')
plt.scatter(x,y)
plt.title('customer distribution')
plt.xlabel('x position')
plt.ylabel('y position')
plt.show()

area = new_x_pos*new_y_pos
print(f'new_base_station_range = {area}')

new_x = []
new_y = []
for t in range(170):
    if x[t] >=new_x_pos-new_y_pos and x[t] <= area:
        new_x.append(x[t])
    if y[t] >= new_x_pos-new_y_pos and y[t] <= area:
        new_y.append(y[t])
print(f'new_x_location = {new_x}')
print(f'new_y_location = {new_y}')

xlen = len(new_x)
ylen = len(new_y)
print(f'{xlen},{ylen}')

if xlen > ylen:
    gap = xlen-ylen
    for f in range(gap):
        new_y.append(0)
else:
    gap = ylen-xlen
    for f in range(gap):
        new_x.append(0)
print(f'gap = {gap}')

print(f'final_new_x_location = {new_x}')
print(f'final_new_y_location = {new_y}')


plt.scatter(new_x,new_y)
plt.title('customer newer distribution')
plt.xlabel('x postion')
plt.ylabel('y position')
plt.show()














